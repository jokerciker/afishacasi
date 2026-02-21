import os
import logging
import html
from datetime import date, timedelta, datetime, time
from zoneinfo import ZoneInfo

from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import Message, KeyboardButton
from aiogram.utils.keyboard import ReplyKeyboardBuilder
from aiogram import F
from aiogram.enums import ParseMode

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

import openpyxl
from dotenv import load_dotenv

# Для вебхуков
import uvicorn
from starlette.applications import Starlette
from starlette.responses import Response, PlainTextResponse
from starlette.routing import Route
from starlette.requests import Request

import database as db

load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_IDS = list(map(int, os.getenv("ADMIN_IDS", "").split(',')))
TIMEZONE = os.getenv("TIMEZONE", "Europe/Moscow")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

db.init_db()

# ---------- Вспомогательные функции ----------
def get_russian_weekday(dt: date) -> str:
    weekdays = {
        0: "пн", 1: "вт", 2: "ср", 3: "чт", 4: "пт", 5: "сб", 6: "вс"
    }
    return weekdays[dt.weekday()]

def format_date_with_weekday(dt: date) -> str:
    return f"<b>{dt.strftime('%d.%m')} ({get_russian_weekday(dt)})</b>"

def format_time(value):
    if value is None:
        return None
    if isinstance(value, time):
        return value.strftime('%H:%M')
    if isinstance(value, datetime):
        return value.strftime('%H:%M')
    if isinstance(value, str):
        return value
    return str(value)

def format_description_with_bold(text: str) -> str:
    """
    Форматирует описание:
    - Экранирует HTML-спецсимволы.
    - Заменяет | на перевод строки.
    - В каждой строке делает жирным текст до первого двоеточия (включая двоеточие).
    """
    if not text:
        return text

    # Экранируем HTML-спецсимволы, чтобы не ломать разметку
    text = html.escape(text)

    text = text.replace('|', '\n')
    lines = text.split('\n')
    formatted_lines = []

    for line in lines:
        line = line.strip()
        if not line:
            formatted_lines.append('')
            continue

        if ':' in line:
            parts = line.split(':', 1)
            key = parts[0].strip()
            value = parts[1].strip()
            formatted_lines.append(f"<b>{key}:</b> {value}")
        else:
            formatted_lines.append(line)

    return '\n'.join(formatted_lines)

def group_events_by_date(events):
    """Группирует события по дате начала (start_date). Возвращает словарь {date_str: [events]}."""
    grouped = {}
    for ev in events:
        start_date = ev[0]  # 'YYYY-MM-DD'
        if start_date not in grouped:
            grouped[start_date] = []
        grouped[start_date].append(ev)
    return grouped

# ---------- Клавиатура ----------
def get_main_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.add(KeyboardButton(text="▶️ Запуск"))
    builder.add(KeyboardButton(text="⏹️ Стоп"))
    builder.row()
    builder.add(KeyboardButton(text="📅 Сегодня"))
    builder.add(KeyboardButton(text="📆 Неделя"))
    builder.adjust(2, 2)
    return builder.as_markup(resize_keyboard=True)

# ---------- Команда /start ----------
@dp.message(Command("start"))
async def cmd_start(message: Message):
    await message.answer(
        "Привет! Я бот с афишей мероприятий.\n"
        "Нажми «▶️ Запуск», чтобы получать ежедневные уведомления в 7 утра о событиях на сегодня и на неделю.\n"
        "Нажми «⏹️ Стоп», чтобы отписаться.",
        reply_markup=get_main_keyboard()
    )

# ---------- Обработка кнопок ----------
@dp.message(F.text == "▶️ Запуск")
async def subscribe(message: Message):
    user_id = message.from_user.id
    username = message.from_user.username
    db.add_user(user_id, username)
    await message.answer("Вы подписаны на утреннюю рассылку! 🎉")
    today = date.today()
    events = db.get_events_for_date(today)
    text = format_events_for_date(events, today)
    await send_long_message(message.chat.id, text, ParseMode.HTML)

@dp.message(F.text == "⏹️ Стоп")
async def unsubscribe(message: Message):
    user_id = message.from_user.id
    db.remove_user(user_id)
    await message.answer("Вы отписались от рассылки. Чтобы вернуться, нажмите «▶️ Запуск».")

@dp.message(F.text == "📅 Сегодня")
async def button_today(message: Message):
    await cmd_today(message)

@dp.message(F.text == "📆 Неделя")
async def button_week(message: Message):
    await cmd_week(message)

# ---------- Команды /today, /week ----------
@dp.message(Command("today"))
async def cmd_today(message: Message):
    today = date.today()
    events = db.get_events_for_date(today)
    text = format_events_for_date(events, today)
    await send_long_message(message.chat.id, text, ParseMode.HTML)

@dp.message(Command("week"))
async def cmd_week(message: Message):
    today = date.today()
    end_of_week = today + timedelta(days=6)
    events = db.get_events_for_week(today, end_of_week)
    if not events:
        await message.answer(f"На ближайшую неделю (с {today.strftime('%d.%m')} по {end_of_week.strftime('%d.%m')}) мероприятий не запланировано.")
        return

    grouped = group_events_by_date(events)
    for day_str in sorted(grouped.keys()):
        day_events = grouped[day_str]
        day_date = datetime.strptime(day_str, '%Y-%m-%d').date()
        day_text = format_events_for_date(day_events, day_date)
        await send_long_message(message.chat.id, day_text, ParseMode.HTML)

# ---------- Команда /clear ----------
@dp.message(Command("clear"))
async def cmd_clear(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("Эта команда только для администратора.")
        return
    db.clear_events()
    await message.answer("✅ Все события удалены из базы данных.")

# ---------- Загрузка Excel ----------
@dp.message(F.document)
async def handle_document(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("У вас нет прав на загрузку файлов.")
        return

    document = message.document
    if not document.file_name.endswith(('.xlsx', '.xls')):
        await message.answer("Пожалуйста, загрузите файл формата Excel (.xlsx или .xls).")
        return

    file_id = document.file_id
    file = await bot.get_file(file_id)
    file_path = file.file_path
    await bot.download_file(file_path, "temp_events.xlsx")

    try:
        wb = openpyxl.load_workbook("temp_events.xlsx", data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        required = ['start_date', 'end_date', 'title']
        if not all(col in headers for col in required):
            await message.answer("Ошибка: файл должен содержать колонки: start_date, end_date, title")
            return

        idx = {h: headers.index(h) for h in headers if h in required + ['time_start', 'time_end', 'location', 'description']}

        events_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue

            def parse_date(cell):
                if isinstance(cell, datetime):
                    return cell.date().isoformat()
                elif isinstance(cell, str):
                    try:
                        return datetime.strptime(cell, '%d.%m.%Y').date().isoformat()
                    except:
                        raise ValueError(f"Неверный формат даты: {cell}")
                else:
                    raise ValueError(f"Неверный тип даты: {cell}")

            try:
                start_date = parse_date(row[idx['start_date']])
                end_date = parse_date(row[idx['end_date']])
            except Exception as e:
                await message.answer(f"Ошибка в строке: {e}")
                continue

            time_start = format_time(row[idx.get('time_start')]) if 'time_start' in idx else None
            time_end = format_time(row[idx.get('time_end')]) if 'time_end' in idx else None
            title = row[idx['title']]
            location = row[idx.get('location')] if 'location' in idx else None
            description = row[idx.get('description')] if 'description' in idx else None

            events_list.append((start_date, end_date, time_start, time_end, title, location, description))

        wb.close()
        db.clear_events()
        db.insert_events(events_list)

        await message.answer(f"Расписание успешно обновлено! Загружено событий: {len(events_list)}")

    except Exception as e:
        await message.answer(f"Произошла ошибка при обработке файла: {e}")
    finally:
        if os.path.exists("temp_events.xlsx"):
            os.remove("temp_events.xlsx")

# ---------- Форматирование сообщений ----------
def format_events_for_date(events, target_date):
    if not events:
        return f"На <b>{target_date.strftime('%d.%m.%Y')}</b> мероприятий не запланировано."

    lines = [f"<b>Афиша на {target_date.strftime('%d.%m.%Y')}:</b>\n"]
    for ev in events:
        start, end, ts, te, title, loc, desc = ev
        start_dt = datetime.strptime(start, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end, '%Y-%m-%d').date()

        if start_dt != end_dt:
            date_str = f"{format_date_with_weekday(start_dt)} – {format_date_with_weekday(end_dt)}"
        else:
            date_str = format_date_with_weekday(start_dt)

        time_str = ""
        if ts and te:
            time_str = f" {ts}–{te}"
        elif ts:
            time_str = f" {ts}"

        # Экранируем title и location
        title_escaped = html.escape(title)
        loc_escaped = html.escape(loc) if loc else None

        line = f"• {date_str}{time_str} – {title_escaped}"
        if loc_escaped:
            line += f" ({loc_escaped})"
        if desc:
            formatted_desc = format_description_with_bold(desc)
            line += f"\n  <i>{formatted_desc}</i>"
        lines.append(line)
        lines.append("")

    return "\n".join(lines)

# ---------- Функция для отправки длинных сообщений ----------
async def send_long_message(chat_id: int, text: str, parse_mode: str = None):
    """Разбивает длинное сообщение на части и отправляет их."""
    MAX_LENGTH = 4096
    while len(text) > MAX_LENGTH:
        # Ищем последний перенос строки в пределах лимита
        split_at = text.rfind('\n', 0, MAX_LENGTH)
        if split_at == -1:
            split_at = MAX_LENGTH
        part = text[:split_at]
        text = text[split_at:].lstrip()
        await bot.send_message(chat_id, part, parse_mode=parse_mode)
    if text:
        await bot.send_message(chat_id, text, parse_mode=parse_mode)

# ---------- Планировщик ----------
scheduler = AsyncIOScheduler(timezone=TIMEZONE)

async def daily_mailing():
    today = date.today()
    users = db.get_active_users()
    for user_id in users:
        try:
            # Сегодня
            today_events = db.get_events_for_date(today)
            today_text = format_events_for_date(today_events, today)
            if today_text:
                await send_long_message(user_id, today_text, ParseMode.HTML)

            # Неделя (по дням)
            end_of_week = today + timedelta(days=6)
            week_events = db.get_events_for_week(today, end_of_week)
            if week_events:
                grouped = group_events_by_date(week_events)
                for day_str in sorted(grouped.keys()):
                    day_events = grouped[day_str]
                    day_date = datetime.strptime(day_str, '%Y-%m-%d').date()
                    day_text = format_events_for_date(day_events, day_date)
                    await send_long_message(user_id, day_text, ParseMode.HTML)
        except Exception as e:
            logging.error(f"Не удалось отправить сообщение пользователю {user_id}: {e}")

scheduler.add_job(daily_mailing, CronTrigger(hour=7, minute=0, timezone=TIMEZONE))

# ---------- ВЕБХУКИ (для Render) ----------
async def on_startup():
    webhook_url = f"{os.environ.get('RENDER_EXTERNAL_URL', '')}/webhook"
    await bot.set_webhook(webhook_url)
    scheduler.start()
    logging.info(f"Бот запущен с вебхуком: {webhook_url}")

async def on_shutdown():
    await bot.delete_webhook()
    logging.info("Бот остановлен")

async def webhook(request: Request) -> Response:
    update_data = await request.json()
    update = types.Update(**update_data)
    await dp.feed_update(bot, update)
    return Response()

async def healthcheck(request: Request) -> PlainTextResponse:
    return PlainTextResponse("OK")

app = Starlette(
    routes=[
        Route("/webhook", webhook, methods=["POST"]),
        Route("/healthcheck", healthcheck, methods=["GET"]),
    ],
    on_startup=[on_startup],
    on_shutdown=[on_shutdown]
)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
